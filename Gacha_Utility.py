import os
import re
import pandas as pd
from openpyxl.styles import NamedStyle, Alignment, Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

SOURCE_DIR = "gachafiles" # The source for the upstream files

def write_to_file(filename, data):
    try:
        with open(filename, 'w') as file:
            entries = data.apply(lambda row: f"{row.name+1}. {row["Name"]},{row["Rarity"]}\n#{row["Description"]}\n", axis=1)
            for entry in entries:
                file.write(entry)
    except FileNotFoundError as e:
        messagebox.showerror("Error", "File not found: {e}")
    except PermissionError as e:
        messagebox.showerror("Error", "Permission denied.\nCould not access file: {e}.")

def read_from_file(filename):
    data = []
    path = os.path.join(SOURCE_DIR, filename)
    try:
        with open(path, 'r') as file:
            for line in file:
                numberPrefix = r'^(\d+\.\s*)'
                match = re.match(numberPrefix, line) # Checks if the line starts with a number: "23." or "167. "
                if match:
                    entry = {}
                    parts = re.sub(numberPrefix, '', line).strip().split(',') # Gets parts without heading
                    entry["Name"] = parts[0]
                    entry["Rarity"] = float(parts[1])
                    data.append(entry)
                else:
                    if "Description" not in data[-1]:
                        poundPrefix = r'^(#\s*)'
                        desc = re.sub(poundPrefix, '', line) # Set description while removing the # prefix
                        data[-1]["Description"] = desc.rstrip()
                    else:
                        # There aren't currently any multi-line descriptions, so if this triggers
                        # then either that has changed, or there's an error somewhere.
                        print(f"Description found while non-heading line detected for entry '{data[-1]["Name"]}' in {filename}")
        return data
    except FileNotFoundError as e:
        messagebox.showerror("Error", "File not found: {e}")
    except PermissionError as e:
        messagebox.showerror("Error", "Permission denied.\nCould not access file: {e}.")
    
def write_to_spreadsheet(filename, data):
    try:
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            wb = writer.book

            # Cell formatting
            header = NamedStyle(name="header")
            header.font = Font(bold=True)
            header.alignment = Alignment(horizontal="center", vertical="center")
            wb.add_named_style(header)

            name = NamedStyle(name="name")
            name.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            wb.add_named_style(name)

            rarity = NamedStyle(name="rarity")
            rarity.alignment = Alignment(horizontal="center", vertical="center")
            rarity.number_format = "0.0"
            wb.add_named_style(rarity)

            description = NamedStyle(name="description")
            description.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            wb.add_named_style(description)

            # Conditional formatting for Rarities
            conditionalFormatting = {
                "trash": CellIsRule(operator="between",
                                formula=["0.1", "0.9"],
                                fill=PatternFill(start_color="F9CB9C", end_color="F9CB9C", fill_type="solid")), # Hex code must use capital letters
                "common": CellIsRule(operator="between",
                                formula=["1.0", "1.9"],
                                fill=PatternFill(start_color="74573E", end_color="74573E", fill_type="solid")),
                "uncommon": CellIsRule(operator="between",
                                formula=["2.0", "2.9"],
                                fill=PatternFill(start_color="ACB9B8", end_color="ACB9B8", fill_type="solid")),
                "rare": CellIsRule(operator="between",
                                formula=["3.0", "3.9"],
                                fill=PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")),
                "elite": CellIsRule(operator="between",
                                formula=["4.0", "4.9"],
                                fill=PatternFill(start_color="4A86E8", end_color="4A86E8", fill_type="solid")),
                "epic": CellIsRule(operator="between",
                                formula=["5.0", "5.9"],
                                fill=PatternFill(start_color="9900FF", end_color="9900FF", fill_type="solid")),
                "legendary": CellIsRule(operator="between",
                                formula=["6.0", "6.9"],
                                fill=PatternFill(start_color="F1C232", end_color="F1C232", fill_type="solid")),
                "mythical": CellIsRule(operator="between",
                                formula=["7.0", "7.9"],
                                fill=PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")),
                "divine": CellIsRule(operator="between",
                                formula=["8.0", "8.9"],
                                fill=PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")),
                "transcendent": CellIsRule(operator="between",
                                formula=["9.0", "9.9"],
                                fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"))
            }
            for cat in data:
                df = pd.DataFrame(data[cat])
                df.to_excel(writer, sheet_name=cat, header=False, index=False, startrow=1) # Skip header row
                ws = writer.sheets[cat]

                # Set header row now
                for i, value in enumerate(df.columns.values):
                    ws.cell(row=1, column=i+1, value=value)
                    ws.cell(row=1, column=i+1).style="header"

                # Formatting Name Column
                for col in ws.iter_cols(min_col=1, max_col=1, min_row=2):
                    for cell in col:
                        if cell.value is not None:
                            cell.style = "name"
                
                # Formatting Rarity Column
                for col in ws.iter_cols(min_col=2, max_col=2, min_row=2):
                    for cell in col:
                        if cell.value is not None:
                            cell.style = "rarity"

                # Add conditional formatting to Rarity Column
                for format in conditionalFormatting:
                    ws.conditional_formatting.add("B2:B1048576", conditionalFormatting[format])
                
                # Formatting Description Column
                for col in ws.iter_cols(min_col=3, max_col=3, min_row=2):
                    for cell in col:
                        if cell.value is not None:
                            cell.style = "description"
                
                # Set column widths
                ws.column_dimensions["A"].width = 30
                ws.column_dimensions["B"].width = 10
                ws.column_dimensions["C"].width = 85
        messagebox.showinfo("Success", f"File was successfully saved: {filename}")
    except PermissionError as e:
        messagebox.showerror("Error", f"Permission denied.\nCould not write file: {e}")
    except Exception as e:
        messagebox.showerror("Error", f"An unexpected error occured: {e}")

def read_from_spreadsheet(filename, output_folder):
    try:
        with pd.ExcelFile(filename) as xls:
            for sheet in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet)
                output_file = os.path.join(output_folder, sheet+".txt")
                write_to_file(output_file, df)
        messagebox.showinfo("Success", f"Files have been successfully saved at {output_folder}")
    except PermissionError as e:
        messagebox.showerror("Error", f"Permission denied.\nCould not write file: {e}")
    except Exception as e:
        messagebox.showerror("Error", f"An unexpected error occured: {e}")

def convert_to_spreadsheet():
    if os.path.exists(SOURCE_DIR):
        files = []
        data = {}
        for file in os.listdir(SOURCE_DIR):
            if file.endswith(".txt"):
                files.append(file)
        if files:
            ext = r'(\.txt)$'
            for file in files:
                key = re.sub(ext, '', file)
                data[key] = read_from_file(file)

            output_file = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                title="Save As",
                                                filetypes=(("Excel Workbook", "*.xlsx"),),
                                                initialdir=".",
                                                initialfile="gacha")
            if output_file:
                write_to_spreadsheet(output_file, data)
        else:
            print(f"No text files found in {SOURCE_DIR}.")
    else:
        print(f"Source directory not found: {SOURCE_DIR}")

def convert_from_spreadsheet():
    filename = filedialog.askopenfilename(defaultextension=".xlsx",
                                    filetypes=(("Excel files", "*.xlsx"),),
                                    title="Open File")
    if filename:
        output_folder = filedialog.askdirectory(title="Select Folder")

        while True:
            if output_folder:
                if os.listdir(output_folder):
                    confirm = messagebox.askyesno(
                        "Directory Not Empty",
                        f"The directory '{output_folder}' is not empty. Selecting this folder may overwrite files.\n\nAre you sure you want to use it?"
                    )
                    if confirm:
                        break
                    else:
                        output_folder = filedialog.askdirectory(title="Select Folder", initialdir=".")
                else:
                    # Directory is empty
                    break
            else:
                # User did not select directory
                return None
        read_from_spreadsheet(filename, output_folder)

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Gacha Utility")
    root.resizable(False, False)
    root.configure(bg="#e6e6e6")
    root.geometry("300x200")

    frame = tk.Frame(root, bg=root["bg"])
    frame.place(relx=0.5, rely=0.5, anchor="center")

    style = ttk.Style()
    style.theme_use("clam")
    style.configure(
        "TButton",
        font=(("Segoe UI", "Helvetica", "Arial"), 12, "bold"),
        background="#5a4bad",
        foreground="white",
        padding=(12, 12),
        relief="flat",
        focusthickness=0,
        focuscolor="none"
    )

    style.map(
        "TButton",
        background=[
            ("pressed", "#6657bb"),
            ("active", "#5244a0")
            
        ],
        foreground=[
            ("disabled", "#AAAAAA")
        ]
    )

    convertToButton = ttk.Button(frame, text="Convert Data to Spreadsheet", command=convert_to_spreadsheet, style="TButton")
    convertToButton.pack(pady=10, expand=True, fill="x")

    convertFromButton = ttk.Button(frame, text="Convert Data from Spreadsheet", command=convert_from_spreadsheet, style="TButton")
    convertFromButton.pack(pady=10, expand=True, fill="x")

    root.mainloop()